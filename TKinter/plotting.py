import tkinter as tk
from tkinter import ttk
import paho.mqtt.publish as publish
import paho.mqtt.client as mqtt
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from screeninfo import get_monitors
import pandas as pd
import time
import tkinter.filedialog as filedialog
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font  # Agregar esta línea para importar Font


t0 = time.monotonic()
# Configuración MQTT
mqtt_host = "tonivivescabaleiro.com"
mqtt_port = 1883
mqtt_topic_odrive = "odrive"
mqtt_topic_data = "data"
 
# Función para enviar el texto por MQTT
def enviar_a_odrive():
    texto = float(texto_input.get()) * 58/43
    print("texto = ",texto)
    text_to_send = str(texto)
    print("text to send = ",text_to_send)
    publish.single(mqtt_topic_odrive, text_to_send, hostname=mqtt_host, port=mqtt_port, qos=2, retain=True)
    # Agrega el instante de tiempo actual a la lista
    time_instant.append(timestamps[-1])

      # Programar el envío de 0 después del tiempo especificado
    #root.after(int(tiempo_input.get()) * 1000, enviar_cero)

def enviar_cero():
    publish.single(mqtt_topic_odrive, "0", hostname=mqtt_host, port=mqtt_port, qos=2, retain=True)

def cerrar_aplicacion():
    descargar_datos()
    root.quit()



# Función para descargar datos (agrega tu funcionalidad)
# def descargar_datos():
#     try:
#         data = {
#             "Timestamp": timestamps,
#             "Position": positions,
#             "Intensity": intensities,
#             "Voltage": voltages,
#             "buttonPressed": time_instant
#         }
        
#         # Permite al usuario seleccionar la ubicación y nombre del archivo
#         file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        
#         if file_path:
#             df = pd.DataFrame(data)
#             df.to_excel(file_path, index=False, engine="openpyxl")
#             print(f"Datos guardados en: {file_path}")
#         else:
#             print("Operación de guardado cancelada.")
#     except Exception as e:
#         print(f"Error al guardar los datos: {str(e)}")

# def descargar_datos():
#     print(time_instant)
#     try:
#         data = {
#             "Timestamp": timestamps,
#             "Position": positions,
#             "Intensity": intensities,
#             "Voltage": voltages,
#             "Times Instants": times_instants
#         }

#         # Permite al usuario seleccionar la ubicación y nombre del archivo
#         file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

#         if file_path:
#             df = pd.DataFrame(data)
#             df
#             df.to_excel(file_path, index=False, engine="openpyxl")
            
#             # Cargar el archivo Excel creado para modificarlo
#             book = load_workbook(file_path)
#             writer = pd.ExcelWriter(file_path, engine='openpyxl')
#             writer.book = book

#             # Seleccionar la hoja en la que deseas trabajar
#             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#             sheet = book.active

#             # Buscar y poner en negrita los valores más cercanos
#             array_indices = [1 if value in time_instant else 0 for value in timestamps]

#             for value in time_instant:
#                 closest_timestamp = min(timestamps, key=lambda x: abs(x - value))
#                 cell = sheet.cell(row=timestamps.index(closest_timestamp) + 2, column=1)
#                 cell.font = Font(bold=True)
#             print(array_indices)


#             writer.save()
#             writer.close()
#             print(f"Datos guardados en: {file_path}")
#         else:
#             print("Operación de guardado cancelada.")
#     except Exception as e:
#         print(f"Error al guardar los datos: {str(e)}")

# def descargar_datos():
#     try:
#         data = {
#             "Timestamp": timestamps,
#             "Position": positions,
#             "Intensity": intensities,
#             "Voltage": voltages,
#             "Times Instants": times_instants
#         }

#         # Permite al usuario seleccionar la ubicación y nombre del archivo
#         file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

#         if file_path:
#             df = pd.DataFrame(data)
#             df.to_excel(file_path, index=False, engine="openpyxl")

#             # Cargar el archivo Excel creado para modificarlo
#             book = load_workbook(file_path)
#             writer = pd.ExcelWriter(file_path, engine='openpyxl')
#             writer.book = book

#             # Seleccionar la hoja en la que deseas trabajar
#             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#             sheet = book.active

#             # Buscar y poner en negrita los valores más cercanos
#             for value in time_instant:
#                 closest_timestamp = min(timestamps, key=lambda x: abs(x - value))
#                 index = timestamps.index(closest_timestamp) + 2      
#                 cell = sheet.cell(row=index, column=5)
#                 cell.value = value              
#                 # Reemplazar el valor en times_instants
#                 times_instants[index - 2] = value
#                 print(time)
#             writer.save()
#             writer.close()
#             print(f"Datos guardados en: {file_path}")
#         else:
#             print("Operación de guardado cancelada.")
#     except Exception as e:
#         print(f"Error al guardar los datos: {str(e)}")# Función de callback cuando se recibe un mensaje MQTT en el tópico "data"

def descargar_datos():
    
    try:
        for value in time_instant:
            closest_timestamp = min(timestamps, key=lambda x: abs(x - value))
            index = timestamps.index(closest_timestamp) + 2      
            times_instants[index - 2] = value
        data = {
            "Timestamp": timestamps,
            "Position": positions,
            "Intensity": intensities,
            "Voltage": voltages,
            "Times Instants": times_instants
        }

        # Permite al usuario seleccionar la ubicación y nombre del archivo
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if file_path:
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, engine="openpyxl")

            print(f"Datos guardados en: {file_path}")
        else:
            print("Operación de guardado cancelada.")
    except Exception as e:
        print(f"Error al guardar los datos: {str(e)}")

def on_message(client, userdata, msg):
    try:
        data = msg.payload.decode("utf-8")
        data_dict = eval(data)
        if isinstance(data_dict, dict) and "timestamp" in data_dict and "position" in data_dict and "intensity" in data_dict and "voltage" in data_dict and "torque" in data_dict :
            timestamps.append(data_dict["timestamp"])
            positions.append(data_dict["position"]*43/58)
            motor_positions.append(data_dict["position"])
            intensities.append(data_dict["intensity"])
            voltages.append(data_dict["voltage"])
            torques.append(data_dict["torque"])
            times_instants.append(0)
            # Actualiza los gráficos en tiempo real
            actualizar_graficos()
    except Exception as e:
        print("Error al procesar mensaje MQTT:", str(e))

# Función para actualizar los gráficos en tiempo real
def actualizar_graficos():
    ax_posicion.clear()
    ax_posicion.plot(timestamps, positions, color='b')
    ax_posicion.plot(timestamps, motor_positions, color='r')
    ax_posicion.set_xlabel("Tiempo")
    ax_posicion.set_ylabel("Posición")
    ax_posicion.set_title("Posición en Tiempo Real")
    
    ax_intensidad.clear()
    ax_intensidad.plot(timestamps, intensities)
    ax_intensidad.set_xlabel("Tiempo")
    ax_intensidad.set_ylabel("Intensidad")
    ax_intensidad.set_title("Intensidad en Tiempo Real")
    
    ax_voltaje.clear()
    ax_voltaje.plot(timestamps, voltages)
    ax_voltaje.set_xlabel("Tiempo")
    ax_voltaje.set_ylabel("Voltaje")
    ax_voltaje.set_title("Voltaje en Tiempo Real")
    
    ax_torque.clear()
    ax_torque.plot(timestamps, torques)
    # Agrega líneas verticales en el gráfico de torque en los instantes de tiempo almacenados en la lista 'time_instant'
    for instant in time_instant:
        ax_torque.axvline(x=instant, color='r', linestyle='--')
   
    ax_torque.set_xlabel("Tiempo")
    ax_torque.set_ylabel("Torque")
    ax_torque.set_title("Torque en Tiempo Real")
    
    canvas.draw()

# Configuración del cliente MQTT
mqtt_client = mqtt.Client("tkinter_mqtt_client")
mqtt_client.on_message = on_message
mqtt_client.connect(mqtt_host, mqtt_port)
mqtt_client.subscribe(mqtt_topic_data)
mqtt_client.loop_start()

# Creación de la interfaz de usuario con Tkinter
root = tk.Tk()
root.title("Envío a ODrive y Gráficos en Tiempo Real")

# Hacer que la ventana sea pantalla completa

#root.attributes('-fullscreen', True)

# Obtiene el tamaño de la pantalla
monitors = get_monitors()
if len(monitors) > 0:
    monitor = monitors[0]  # Utiliza el primer monitor encontrado
    screen_width = monitor.width
    screen_height = monitor.height
else:
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

# Ajusta el tamaño de la figura de Matplotlib para que coincida con el tamaño de la pantalla
fig, axs = plt.subplots(2, 2, figsize=(screen_width / 100, screen_height / 100), sharex=True)
fig.subplots_adjust(hspace=0.2, wspace=0.5, bottom=0.2)  # Ajusta el espacio horizontal con wspace

ax_posicion = axs[0, 0]
ax_intensidad = axs[0, 1]
ax_voltaje = axs[1, 0]
ax_torque = axs[1, 1]

timestamps = []
positions = []
intensities = []
voltages = []
torques = []
finalData = []
time_instant = []
times_instants = []
motor_positions = []
# Integrar los gráficos en la ventana de Tkinter usando FigureCanvasTkAg
canvas = FigureCanvasTkAgg(fig, master=root)
canvas.get_tk_widget().grid(row=1, column=0)

# Etiqueta y entrada de texto para el envío MQTT
frame_top = ttk.Frame(root)
frame_top.grid(row=0, column=0, padx=10, pady=10)

texto_label = ttk.Label(frame_top, text="Texto a enviar a ODrive:")
texto_label.grid(row=0, column=0, padx=5, pady=5, sticky="E")

texto_input = ttk.Entry(frame_top)
texto_input.grid(row=1, column=0, padx=5, pady=5)

# Nueva entrada de texto para el tiempo
tiempo_label = ttk.Label(frame_top, text="Tiempo de espera (segundos):")
tiempo_label.grid(row=0, column=1, padx=5, pady=5, sticky="E")

tiempo_input = ttk.Entry(frame_top)
tiempo_input.grid(row=1, column=1, padx=5, pady=5)

# Botones para enviar y descargar datos
frame_buttons = ttk.Frame(root)
frame_buttons.grid(row=0, column=0, padx=10, pady=10)

enviar_button = ttk.Button(frame_top, text="Enviar a ODrive", command=enviar_a_odrive)
enviar_button.grid(row=0, column=3, padx=5, pady=5)

cerrar_button = ttk.Button(frame_top, text="save and close", command=cerrar_aplicacion)
cerrar_button.grid(row=0, column=5, padx=5, pady=5, sticky="NE")


# Inicia la interfaz gráfica
root.mainloop()
